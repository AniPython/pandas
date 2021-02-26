from openpyxl.styles import (
    Alignment,  # 对齐
)
from openpyxl import load_workbook

wb = load_workbook('tips.xlsx')
ws = wb.active

ws['A1'].alignment = Alignment(
    horizontal='center',  # {'fill', 'right', 'general', 'left', 'justify', 'centerContinuous', 'distributed', 'center'}
    vertical='bottom',  # {'bottom', 'justify', 'distributed', 'top', 'center'}
    wrap_text=True
)

ws['B1'].alignment = Alignment(
    horizontal='right',  # {'fill', 'right', 'general', 'left', 'justify', 'centerContinuous', 'distributed', 'center'}
    vertical='top',  # {'bottom', 'justify', 'distributed', 'top', 'center'}
    wrap_text=False
)

wb.save('tips_4.xlsx')