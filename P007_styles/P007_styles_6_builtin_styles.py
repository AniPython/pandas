from openpyxl import load_workbook


wb = load_workbook('tips.xlsx')
ws = wb.active

for c in ws[1]:
    c.style = "Pandas"

for c in ws['B'][1:] + ws['E'][1:]:
    c.style = "Good"

wb.save('builtin.xlsx')
