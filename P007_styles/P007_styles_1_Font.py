from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook('tips.xlsx')
ws = wb.active

font = Font(
    name='Apple SD Gothic Neo',    # 字体名称
    size=30,         # 字体大小
    bold=True,      # 是否黑体
    italic=True,    # 是否斜体
    underline='doubleAccounting',  # 下划线 {'double', 'single', 'singleAccounting', 'doubleAccounting'}
    color='AAFF0000'   # 颜色
)
# ws['A1'].font = font

for c in ws[1]:
    c.font = font

wb.save('tips_1.xlsx')