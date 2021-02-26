from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string

wb = load_workbook('tips.xlsx')
ws = wb.active

# ws.column_dimensions['B'].width = 20
# ws.row_dimensions[2].height = 60

# print(get_column_letter(1))
# print(get_column_letter(2))
# print(get_column_letter(3))

# print(column_index_from_string('A'))
# print(column_index_from_string('B'))
# print(column_index_from_string('C'))

# print(ws.max_column)
# print(ws.max_row)

for i in range(ws.max_row):
    # print(i)
    ws.row_dimensions[i + 1].height = 60

for i in range(ws.max_column):
    # print(i)
    letter = get_column_letter(i + 1)
    ws.column_dimensions[letter].width = 20

wb.save('tips_5.xlsx')
