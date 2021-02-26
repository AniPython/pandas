from openpyxl.styles import (
    PatternFill,  # 颜色填充
    GradientFill,  # 渐变填充
)
from openpyxl import load_workbook

wb = load_workbook('tips.xlsx')
ws = wb.active

# {'lightGrid', 'lightVertical', 'lightTrellis',
# 'darkDown', 'lightDown', 'mediumGray',
# 'gray125', 'darkHorizontal', 'gray0625',
# 'darkGray', 'lightUp', 'darkGrid',
# 'darkUp', 'lightGray', 'darkTrellis',
# 'solid', 'lightHorizontal', 'darkVertical'}
ws['A1'].fill = PatternFill(
    patternType='solid',
    fgColor='FDEB71'
)
ws['B1'].fill = PatternFill(
    patternType='lightGrid',
    fgColor='FDEB71'
)

ws['C1'].fill = GradientFill(
    degree=0,
    stop=('CE9FFC', '7367F0')
)
ws['D1'].fill = GradientFill(
    degree=45,
    stop=('CE9FFC', '7367F0')
)
ws['E1'].fill = GradientFill(
    degree=90,
    stop=('CE9FFC', '7367F0')
)

for c in ws['B']:
    c.fill = PatternFill(
        patternType='solid',
        fgColor='F55555'
    )

wb.save('tips_2.xlsx')
