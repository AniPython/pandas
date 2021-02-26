"""
openpyxl 与 pandas 相互转化
"""
from openpyxl import Workbook, load_workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

"""
pandas -> openpyxl
"""
df = pd.DataFrame(
    [[1, 2, 3],
     [4, 5, 6],
     [7, 8, 9]],
    columns=list('ABC'),
    index=list('XYZ')
)
# print(df)

rows = dataframe_to_rows(df, index=False, header=True)
print(rows)

wb = Workbook()
ws = wb.active

for row in rows:
    ws.append(row)

wb.save('1.xlsx')

"""
openpyxl -> pandas
"""
wb = load_workbook('tips.xlsx')
ws = wb.active

data = list(ws.values)
# print(data)
df = pd.DataFrame(data[1:], columns=data[0])

print(df.head())