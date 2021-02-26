from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os


def insert_img_to_excel(
        filename,
        by_col,
        to_col,
        img_folder
):
    """
    插入图片到 excel
    :param filename: 文件路径
    :param by_col: 依靠列 (A, B, C ...)
    :param to_col: 插入到列 (A, B, C ...)
    :param img_folder: 图片源的文件夹
    :return: None
    """
    wb = load_workbook(filename)
    ws = wb.active
    for i, c in enumerate(ws[by_col], start=1):
        # print(i, c.value)
        img_ffn = os.path.join(img_folder, c.value + '.jpg')
        print(i, img_ffn)
        try:
            ws.add_image(
                img=Image(img_ffn),
                anchor=to_col + str(i)
            )
        except:
            print(c.value, '匹配不到图片')
    wb.save(filename)


if __name__ == '__main__':
    insert_img_to_excel(
        '1.xlsx', 'A', 'B',
        img_folder="/Users/Yi/Mirror/我的python教程/Pandas办公自动化/P006_add_image/img"
    )

