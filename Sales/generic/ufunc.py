import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


# 上下合并全部表格
def v_concat_tables(folder) -> pd.DataFrame:
    df_list = []
    for fn in os.listdir(folder):
        afn = os.path.join(folder, fn)
        print(afn)
        df_temp = pd.read_excel(afn)
        df_list.append(df_temp)
    df = pd.concat(df_list, ignore_index=True)
    print(df.info())
    return df


def insert_img_to_excel(
        filename,
        by_col,
        to_col,
        img_folder
):
    """
    插入图片到 excel
    :param filename: 文件路径
    :param by_col: 依靠列 (A, B, C ...)
    :param to_col: 插入到列 (A, B, C ...)
    :param img_folder: 图片源的文件夹
    :return: None
    """
    wb = load_workbook(filename)
    ws = wb.active
    for i, c in enumerate(ws[by_col], start=1):
        # print(i, c.value)
        img_ffn = os.path.join(img_folder, c.value + '.jpg')
        print(i, img_ffn)
        try:
            img = Image(img_ffn)
            img.height = 80
            img.width = 80
            ws.add_image(
                img=img,
                anchor=to_col + str(i)
            )
        except:
            print(c.value, '匹配不到图片')
    wb.save(filename)



if __name__ == '__main__':
    df_tb = v_concat_tables('/Users/Yi/Mirror/我的python教程/Sales/data_source/tb_sales')
    # df_tb.to_excel('tb.xlsx')